from __future__ import annotations

import json
import re
import subprocess
from dataclasses import dataclass, asdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"E:\Roadmaps\phanmanhcuongdev\personal-dossier-builder")
ROADMAPS = Path(r"E:\Roadmaps")
LAP = Path(r"E:\Lap")

OUT = ROOT / "extracted"
NOTES = ROOT / "working-notes"
OUT.mkdir(parents=True, exist_ok=True)
NOTES.mkdir(parents=True, exist_ok=True)

IGNORE_DIRS = {
    ".git", ".venv", "venv", "env", "node_modules", "build", "target", "obj",
    "bin", ".gradle", ".idea", ".vs", "dist", ".next", "__pycache__",
}
ALLOWED_EXT = {".md", ".txt", ".docx", ".xlsx", ".json", ".yml", ".yaml", ".xml"}
IMPORTANT_NAMES = {
    "readme.md", "readme_llm.md", "tutorial.md", "integration.md", "pom.xml",
    "package.json", "docker-compose.yml", "docker-compose.yaml", "docker-image.yml",
    "envswitch.csproj", "cmakelists.txt", "roadmap.md",
}
SENSITIVE_RE = re.compile(
    r"(password\s*=|token\s*=|api[_-]?key|secret|private key|BEGIN RSA PRIVATE KEY|\.env|credential)",
    re.IGNORECASE,
)


@dataclass
class Source:
    path: str
    type: str
    short_description: str
    relevance: str
    risk_sensitivity: str
    use_in_final: str
    headings: list[str]


def is_ignored(path: Path) -> bool:
    return any(part.lower() in IGNORE_DIRS for part in path.parts)


def should_include(path: Path, base: Path) -> bool:
    if is_ignored(path):
        return False
    suffix = path.suffix.lower()
    if suffix not in ALLOWED_EXT:
        return False
    if base == ROADMAPS:
        return suffix in {".md", ".txt", ".docx", ".xlsx"}
    name = path.name.lower()
    if name in IMPORTANT_NAMES:
        return True
    text_path = str(path).lower()
    return any(key in text_path for key in [
        "student-feedback-system", "ai-translation-wrapper", "window_ui",
        "iot_va_ung_dung", "android", "ttcs",
    ]) and suffix in {".md", ".txt", ".json", ".yml", ".yaml", ".xml"}


def read_docx(path: Path) -> str:
    try:
        result = subprocess.run(
            ["pandoc", str(path), "-t", "markdown", "--wrap=none"],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        return result.stdout
    except Exception as exc:
        return f"[DOCX READ ERROR] {exc}"


def read_xlsx(path: Path) -> str:
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
        lines = [f"Workbook sheets: {', '.join(wb.sheetnames)}"]
        for ws in wb.worksheets:
            headers = [ws.cell(1, col).value for col in range(1, min(ws.max_column, 20) + 1)]
            lines.append(f"{ws.title}: headers={headers}")
        return "\n".join(lines)
    except Exception as exc:
        return f"[XLSX READ ERROR] {exc}"


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return f"[TEXT READ ERROR] {exc}"


def read_source(path: Path) -> str:
    if path.suffix.lower() == ".docx":
        return read_docx(path)
    if path.suffix.lower() == ".xlsx":
        return read_xlsx(path)
    return read_text(path)


def headings(text: str) -> list[str]:
    found = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            found.append(stripped[:160])
        if len(found) >= 12:
            break
    return found


def describe(path: Path, text: str) -> str:
    hs = headings(text)
    if hs:
        return hs[0].lstrip("# ").strip()
    for line in text.splitlines():
        line = line.strip()
        if line and not line.startswith("```"):
            return line[:140]
    return "No readable summary found."


def relevance(path: Path) -> str:
    p = str(path).lower()
    if "12-month-devsecops-roadmap" in p or "roadmap-to-devsecops" in p or "\\docx\\" in p:
        return "Core roadmap / operating model reference"
    if "trackers" in p:
        return "Evidence tracking system reference"
    if "student-feedback-system" in p:
        return "Primary software project evidence"
    if "ai-translation-wrapper" in p:
        return "Worker / RabbitMQ integration evidence"
    if "window_ui" in p:
        return "Local tooling / automation evidence"
    if "iot_va_ung_dung" in p or "android" in p:
        return "School / adjacent project context"
    return "Supporting context"


def use_in_final(path: Path, risk: str) -> str:
    p = str(path).lower()
    if "node_modules" in p or risk.startswith("High"):
        return "No, inventory only"
    if any(key in p for key in ["roadmap", "student-feedback-system", "ai-translation-wrapper", "window_ui", "trackers", "docx"]):
        return "Yes, summarized"
    return "Maybe, supporting context only"


def risk_for(text: str, path: Path) -> str:
    p = str(path).lower()
    if ".env" in p:
        return "High - env file path"
    if SENSITIVE_RE.search(text):
        return "Medium - contains credential/template/security terms; summarize only"
    return "Low"


def collect_sources() -> list[Source]:
    paths = []
    for base in [ROADMAPS, LAP]:
        for path in base.rglob("*"):
            if path.is_file() and should_include(path, base):
                paths.append(path)
    sources: list[Source] = []
    for path in sorted(set(paths), key=lambda p: str(p).lower()):
        text = read_source(path)
        risk = risk_for(text, path)
        src = Source(
            path=str(path),
            type=path.suffix.lower().lstrip(".") or "file",
            short_description=describe(path, text),
            relevance=relevance(path),
            risk_sensitivity=risk,
            use_in_final=use_in_final(path, risk),
            headings=headings(text),
        )
        sources.append(src)
    github = Source(
        path="https://github.com/phanmanhcuongdev",
        type="url",
        short_description="Public GitHub profile for Phan Manh Cuong; profile shows public repositories and visible profile metadata.",
        relevance="Public portfolio presence",
        risk_sensitivity="Low - public profile; avoid personal/social links in dossier",
        use_in_final="Yes, summarized",
        headings=[
            "Public profile: Phan Manh Cuong / phanmanhcuongdev",
            "Visible public repositories include student-feedback-system, sqlserver-replication-lab, distributed-systems-lab, headscale-infra, hotel-management-system, TRR-PTIT",
        ],
    )
    sources.append(github)
    return sources


def write_inventory(sources: list[Source]) -> None:
    lines = [
        "# Source Inventory",
        "",
        "This inventory was generated from readable local files under `E:\\Roadmaps`, selected relevant files under `E:\\Lap`, and public GitHub profile metadata.",
        "",
        "| Source path / URL | Type | Short description | Relevance | Risk / sensitivity | Use in final dossier? |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for s in sources:
        path = s.path.replace("|", "\\|")
        desc = s.short_description.replace("|", "\\|").replace("\n", " ")
        lines.append(f"| `{path}` | {s.type} | {desc} | {s.relevance} | {s.risk_sensitivity} | {s.use_in_final} |")
    (NOTES / "source-inventory.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    sources = collect_sources()
    (OUT / "sources.json").write_text(
        json.dumps([asdict(s) for s in sources], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_inventory(sources)
    print(f"Indexed {len(sources)} sources")
    print(NOTES / "source-inventory.md")
    print(OUT / "sources.json")


if __name__ == "__main__":
    main()
